"""Converter for XLSX/XLS files."""

import logging
from pathlib import Path
from typing import Optional

from office2md.converters.base_converter import BaseConverter

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class XlsxConverter(BaseConverter):
    """Converter for XLSX/XLS files."""

    def __init__(
        self,
        input_path: str,
        output_path: Optional[str] = None,
        include_all_sheets: bool = True,
        **kwargs
    ):
        """
        Initialize XLSX converter.

        Args:
            input_path: Path to input XLSX file
            output_path: Optional output path
            include_all_sheets: Include all sheets (default: True)
            **kwargs: Additional options (extract_images, embed_images, skip_images)
        """
        super().__init__(input_path, output_path, **kwargs)
        self.include_all_sheets = include_all_sheets

    def convert(self) -> str:
        """Convert XLSX to Markdown."""
        if not OPENPYXL_AVAILABLE:
            raise RuntimeError("openpyxl is not available for XLSX conversion")

        try:
            logger.info("Converting XLSX using openpyxl")
            workbook = load_workbook(self.input_path)

            markdown_lines = []

            # Determine which sheets to process
            sheets = workbook.sheetnames
            if not self.include_all_sheets and sheets:
                sheets = [sheets[0]]

            for sheet_name in sheets:
                worksheet = workbook[sheet_name]

                # Add sheet name as heading
                markdown_lines.append(f"## {sheet_name}")
                markdown_lines.append("")

                # Convert sheet to markdown table
                markdown_lines.append(self._sheet_to_markdown(worksheet))
                markdown_lines.append("")

            return "\n".join(markdown_lines)

        except Exception as e:
            logger.error(f"Error converting XLSX: {e}")
            raise

    def _sheet_to_markdown(self, worksheet) -> str:
        """
        Convert a worksheet to Markdown table format.

        Args:
            worksheet: openpyxl Worksheet object

        Returns:
            Markdown table string
        """
        rows = []

        for i, row in enumerate(worksheet.iter_rows(values_only=True)):
            # Skip empty rows
            if not any(row):
                continue

            # Format cells
            cells = [str(cell) if cell is not None else "" for cell in row]
            rows.append("| " + " | ".join(cells) + " |")

            # Add separator after header row
            if i == 0:
                rows.append("|" + "|".join([" --- " for _ in cells]) + "|")

        return "\n".join(rows) if rows else ""
